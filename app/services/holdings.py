"""持仓与交易：移动加权成本、撤销回滚（重放法）、港股人民币折算。

持仓是交易记录的物化视图。任何插入/删除交易后，对受影响股票按时间顺序重放全部交易，
重算持仓数量与移动加权成本——保证绝对一致，撤销只需删交易再重放。

港股折算口径（用户确认）：
- 港股持仓同时维护原币平均成本与人民币平均成本（每笔买入按交易日汇率折算）。
- 汇率缺失时绝不按 1:1 计算：该港股显示原币数据，人民币成本置 None（missing_fx）。
"""
import io
from datetime import datetime

from app.instruments import get_instrument
from app.models.db import get_conn

# 交易必须字段
_REQUIRED = ("code", "side", "price", "quantity")


def _currency_of(conn, code: str) -> str:
    """该股币种（默认 CNY）。"""
    row = conn.execute("SELECT currency FROM stocks WHERE code=?", (code,)).fetchone()
    return (row["currency"] if row and row["currency"] else "CNY") or "CNY"


def rebuild(code: str, conn) -> dict:
    """按时间顺序重放 code 的全部交易，重建持仓。返回持仓 dict。

    人民币口径：原币金额 × 交易日汇率。某笔买入汇率缺失时人民币成本不可算（置 None）。
    """
    currency = _currency_of(conn, code)
    # 按 id（录入顺序）重放：adjust 交易用微秒时间戳避免 UNIQUE 冲突，字符串排序会把
    # 凌晨微秒时间戳排到当天 09:00 买入之前，导致 adjust 先于买入处理报"数量为 0 或负"。
    rows = conn.execute(
        "SELECT * FROM trades WHERE code=? ORDER BY id", (code,)
    ).fetchall()
    qty = 0.0
    avg_cost = 0.0
    total_buy = 0.0
    avg_cost_cny = 0.0
    total_buy_cny = 0.0
    cny_ok = True
    for t in rows:
        amt = t["amount"] or 0.0
        fee = t["fee"] or 0.0
        if t["side"] == "buy":
            total_buy += amt + fee
            new_qty = qty + t["quantity"]
            avg_cost = (qty * avg_cost + amt + fee) / new_qty if new_qty > 0 else 0.0
            amt_cny = t["amount_cny"]
            if amt_cny is None:
                cny_ok = False
            else:
                # 缺汇率绝不按 1:1：本金已有 amount_cny 时费用也必须同汇率；无汇率则费用不计入并标 missing_fx
                if currency == "CNY":
                    fee_cny = fee
                elif t["fx_rate"]:
                    fee_cny = fee * float(t["fx_rate"])
                else:
                    fee_cny = 0.0
                    cny_ok = False
                total_buy_cny += amt_cny + fee_cny
                avg_cost_cny = (qty * avg_cost_cny + amt_cny + fee_cny) / new_qty if new_qty > 0 else 0.0
            qty = new_qty
        elif t["side"] == "adjust":
            # 成本/股数调整：quantity=股数变化量(delta)，amount=成本变化额。
            # 纯成本调整 delta=0；拆股/送股 delta>0、amount=0 → 总成本不变、每股摊薄。
            old_qty = qty
            new_qty = qty + (t["quantity"] or 0.0)
            if new_qty <= 0:
                raise ValueError(f"调整后数量不能为 0 或负（{code}）")
            total_buy += amt
            avg_cost = (old_qty * avg_cost + amt) / new_qty
            amt_cny = t["amount_cny"]
            if amt_cny is None:
                cny_ok = False
            else:
                avg_cost_cny = (old_qty * avg_cost_cny + amt_cny) / new_qty
                total_buy_cny += amt_cny
            qty = new_qty
        else:  # sell
            qty -= t["quantity"]
            if qty < -1e-9:
                raise ValueError(f"卖出数量超过持仓（{code}）")
    status = "active" if qty > 1e-9 else "closed"
    holding = {
        "code": code,
        "quantity": round(qty, 6),
        "avg_cost": round(avg_cost, 4),
        "avg_cost_cny": round(avg_cost_cny, 4) if cny_ok else None,
        "total_buy": round(total_buy, 2),
        "total_buy_cny": round(total_buy_cny, 2) if cny_ok else None,
        "currency": currency,
        "missing_fx": False if cny_ok else True,
        "status": status,
    }
    conn.execute(
        """INSERT INTO holdings(code, quantity, avg_cost, avg_cost_cny, total_buy, total_buy_cny, currency, status)
           VALUES (:code,:quantity,:avg_cost,:avg_cost_cny,:total_buy,:total_buy_cny,:currency,:status)
           ON CONFLICT(code) DO UPDATE SET
             quantity=excluded.quantity, avg_cost=excluded.avg_cost, avg_cost_cny=excluded.avg_cost_cny,
             total_buy=excluded.total_buy, total_buy_cny=excluded.total_buy_cny,
             currency=excluded.currency, status=excluded.status""",
        holding,
    )
    # 清仓只改持仓状态；原始缓存（日K/财务/估值/分位/汇率）长期保留，再次开仓直接复用。
    # 不调用 _purge_stock_cache（见 OPTIMIZATION_PLAN 第 5 节）。
    return holding


def _market_of(inst) -> str:
    """标的交易所代码（stocks.market）：腾讯/新浪 symbol 前缀 → hk/sh/sz/bj。"""
    sym = inst.symbol()
    return sym[:2] if sym and sym[:2] in ("hk", "sh", "sz", "bj") else "sh"


def _ensure_stock(conn, code: str, name: str | None) -> None:
    if name:
        inst = get_instrument(code)
        conn.execute(
            """INSERT INTO stocks(code, name, market, tag, currency) VALUES(?,?,?,?,?)
               ON CONFLICT(code) DO UPDATE SET name=excluded.name, currency=excluded.currency""",
            (code, name, _market_of(inst), inst.tag, inst.currency),
        )


def parse_holdings_excel(data: bytes) -> tuple[list[dict], list[dict]]:
    """解析「汇总持仓.xlsx」的持仓数据 sheet。

    返回 (可导入项, 跳过明细)；可导入项为 {code, name, price, quantity, fee}，
    价格优先取「单位成本」，缺省回退「最新价」；仅支持 A 股/场内基金。
    """
    from openpyxl import load_workbook

    from app.instruments import get_instrument

    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb["持仓数据"] if "持仓数据" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    def col(name: str):
        return header.index(name) if name in header else None

    idx = {
        "code": col("代码"),
        "name": col("名称"),
        "qty": col("持有数量"),
        "cost": col("单位成本"),
        "price": col("最新价"),
    }
    if idx["code"] is None or idx["qty"] is None:
        raise ValueError("Excel 缺少「代码」或「持有数量」列")

    items, skipped = [], []
    for row in rows[1:]:
        code = str(row[idx["code"]]).strip() if row[idx["code"]] is not None else ""
        name = str(row[idx["name"]]).strip() if idx["name"] is not None and row[idx["name"]] is not None else ""
        try:
            qty = float(row[idx["qty"]])
        except (TypeError, ValueError):
            qty = 0.0
        if not code or qty <= 0:
            continue
        try:
            get_instrument(code).symbol()
        except ValueError:
            skipped.append({"code": code, "name": name, "reason": "非A股/代码格式不支持"})
            continue

        price = None
        if idx["cost"] is not None and row[idx["cost"]] is not None:
            try:
                price = float(row[idx["cost"]])
            except (TypeError, ValueError):
                price = None
        if not price and idx["price"] is not None and row[idx["price"]] is not None:
            try:
                price = float(row[idx["price"]])
            except (TypeError, ValueError):
                price = None
        if not price or price <= 0:
            skipped.append({"code": code, "name": name, "reason": "无有效价格"})
            continue

        items.append({
            "code": code,
            "name": name or None,
            "price": round(price, 4),
            "quantity": float(qty),
            "fee": 0.0,
        })
    return items, skipped


def _compute_cny(currency: str, amount: float, trade_date: str) -> tuple[float | None, float | None]:
    """按交易日汇率折算人民币。返回 (fx_rate, amount_cny)。

    CNY 股：fx_rate=1.0、amount_cny=amount。港股：取交易日汇率，缺失返回 (None, None)。
    """
    if currency == "CNY":
        return 1.0, round(amount, 2)
    try:
        from app.services.fx import ensure_fx_for_date

        rate = ensure_fx_for_date("HKD", trade_date)
    except Exception:  # noqa: BLE001 汇率拉取失败不阻断录入
        rate = None
    if rate is None:
        return None, None
    return round(rate, 6), round(amount * rate, 2)


def record_trade(code, side, price, quantity, fee=0.0, trade_time=None, note=None, name=None,
                 *, side_effects: bool = True) -> dict:
    """录入交易并重放持仓。返回 {trade_id, holding}。

    side_effects=False：只写交易+重放，跳过新股同步 / 组合序列重算 / 当日 AI 打分
    （批量导入末尾统一触发，避免 N 只重复全量同步）。
    """
    if not get_instrument(code).can_trade:
        raise ValueError("指数不可交易")
    for f in _REQUIRED:
        if locals().get(f) is None:
            raise ValueError(f"缺少必填字段: {f}")
    if side not in ("buy", "sell"):
        raise ValueError("side 必须为 buy/sell")
    if price <= 0 or quantity <= 0:
        raise ValueError("价格与数量必须为正")
    trade_time = trade_time or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    amount = round(price * quantity, 4)
    # 汇率计算放在事务外：内部需独立连接写 fx_rate_cache，事务内写会锁库
    currency = get_instrument(code).currency
    fx_rate, amount_cny = _compute_cny(currency, amount, trade_time[:10])
    with get_conn() as conn:
        _ensure_stock(conn, code, name)
        cur = conn.execute(
            """INSERT INTO trades(code, side, price, quantity, amount, fee, trade_time, note, fx_rate, amount_cny)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (code, side, price, quantity, amount, fee, trade_time, note, fx_rate, amount_cny),
        )
        trade_id = cur.lastrowid
        try:
            holding = rebuild(code, conn)
        except ValueError:
            raise
    if side_effects:
        # 新股：先同步数据（否则 AI 打分时无该股缓存因子）
        if _is_new_stock(code):
            _sync_stock_data(code)
        # 开仓/清仓引入新持仓权重 → 重算组合综合 PE/PB 序列缓存
        _rebuild_portfolio()
        # 该日 AI 打分失效并后台自动重打分（尽力而为，失败不影响交易录入）
        _trigger_ai_daily(trade_time[:10])
    return {"trade_id": trade_id, "holding": holding}


def import_holdings_items(items: list[dict], on_progress=None) -> dict:
    """空仓一键导入（同步版，测试/兼容用）：逐只写买入后统一同步/重算/打分。

    生产路径走 job_runners.start_holdings_import（按股扇出子任务）。
    on_progress(done, total, current_label) 可选。
    """
    total = len(items)
    codes: list[str] = []
    for i, it in enumerate(items):
        label = f"{it['code']} {it.get('name') or ''}".strip()
        if on_progress:
            on_progress(i, total, label)
        record_trade(
            code=it["code"], side="buy", price=it["price"], quantity=it["quantity"],
            fee=it.get("fee", 0.0), name=it.get("name"),
            side_effects=False,
        )
        codes.append(it["code"])
        if on_progress:
            on_progress(i + 1, total, label)
    for code in codes:
        if on_progress:
            on_progress(total, total, f"同步 {code}")
        _sync_stock_data(code)
    _rebuild_portfolio()
    score_date = datetime.now().strftime("%Y-%m-%d")
    _trigger_ai_daily(score_date)
    return {"imported": total, "codes": codes, "score_date": score_date}


def adjust_cost(code, amount=0.0, delta_qty=0.0, note=None, trade_time=None, name=None,
                is_dividend=False) -> dict:
    """直接调整持仓：调整成本（amount，正=加成本 负=减成本，如分红除权摊薄）与/或调整股数
    （delta_qty，正=加股 负=减股，如拆股/送股；只加股不改总成本 → 每股成本摊薄）。

    在 trades 表插入一条 adjust 记录（quantity=delta_qty、amount=调整额），重放时保持
    持仓物化视图一致。不生成评分快照（非买卖）。is_dividend=True 表示分红除权，计入累计分红。
    """
    if (amount is None or amount == 0) and (delta_qty is None or delta_qty == 0):
        raise ValueError("调整金额与股数不能同时为 0")
    # 微秒精度：允许同一天多次调整而不撞 UNIQUE(code, trade_time, side, price, quantity)
    trade_time = trade_time or datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
    with get_conn() as conn:
        _ensure_stock(conn, code, name)
        h = conn.execute("SELECT quantity FROM holdings WHERE code=?", (code,)).fetchone()
        if not h or h["quantity"] <= 0:
            raise ValueError("当前无持仓，无法调整")
        currency = _currency_of(conn, code)
        if currency == "CNY":
            fx_rate, amount_cny = 1.0, round(amount or 0.0, 2)
        else:
            fx_rate, amount_cny = _compute_cny(currency, amount or 0.0, trade_time[:10])
        cur = conn.execute(
            """INSERT INTO trades(code, side, price, quantity, amount, fee, trade_time, note, fx_rate, amount_cny, is_dividend)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (code, "adjust", 0.0, delta_qty or 0.0, round(amount or 0.0, 4), 0.0, trade_time, note,
             fx_rate, amount_cny, 1 if is_dividend else 0),
        )
        trade_id = cur.lastrowid
        holding = rebuild(code, conn)
    _rebuild_portfolio()
    _trigger_ai_daily(trade_time[:10])
    return {"trade_id": trade_id, "holding": holding}


def delete_trade(trade_id: int) -> dict:
    """撤销交易（删记录 + 重放）。若删除后持仓非法则拒绝并回滚。"""
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM trades WHERE id=?", (trade_id,)).fetchone()
        if not row:
            raise ValueError(f"交易不存在: {trade_id}")
        conn.execute("DELETE FROM trades WHERE id=?", (trade_id,))
        try:
            holding = rebuild(row["code"], conn)
        except ValueError:
            raise
    _rebuild_portfolio()
    _trigger_ai_daily(row["trade_time"][:10])
    return {"deleted_trade_id": trade_id, "holding": holding}


def update_trade(trade_id: int, **fields) -> dict:
    """修改单笔交易：字段为 None 则保持原值。改后重放持仓并重算涉及日的综合分。

    若改动 code/日期，涉及新旧两日的综合分都重算。
    """
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM trades WHERE id=?", (trade_id,)).fetchone()
        if not row:
            raise ValueError(f"交易不存在: {trade_id}")

        new_code = fields.get("code") or row["code"]
        if new_code != row["code"] and not get_instrument(new_code).can_trade:
            raise ValueError("指数不可交易")
        new = {
            "code": new_code,
            "side": fields.get("side") or row["side"],
            "price": fields.get("price") if fields.get("price") is not None else row["price"],
            "quantity": fields.get("quantity") if fields.get("quantity") is not None else row["quantity"],
            "fee": fields.get("fee") if fields.get("fee") is not None else row["fee"],
            "trade_time": fields.get("trade_time") or row["trade_time"],
            "note": fields["note"] if "note" in fields else row["note"],
        }
        if row["side"] == "adjust":
            # 成本/股数调整：只允许改 note/时间；quantity=拆股增量、amount=成本调整额必须保留
            new["side"] = "adjust"
            new["price"] = 0.0
            new["quantity"] = row["quantity"] or 0.0
            new["fee"] = row["fee"] or 0.0
            amount = row["amount"] or 0.0
        else:
            if new["side"] not in ("buy", "sell"):
                raise ValueError("side 必须为 buy/sell")
            if new["price"] <= 0 or new["quantity"] <= 0:
                raise ValueError("价格与数量必须为正")
            amount = round(new["price"] * new["quantity"], 4)
        currency = _currency_of(conn, new["code"])
    # 汇率计算放在写事务外（内部写 fx_rate_cache，避免与外层写锁冲突）
    fx_rate, amount_cny = _compute_cny(currency, amount, new["trade_time"][:10])
    with get_conn() as conn:
        try:
            conn.execute(
                """UPDATE trades SET code=?, side=?, price=?, quantity=?, amount=?, fee=?, trade_time=?, note=?, fx_rate=?, amount_cny=?
                   WHERE id=?""",
                (new["code"], new["side"], new["price"], new["quantity"], amount,
                 new["fee"], new["trade_time"], new["note"], fx_rate, amount_cny, trade_id),
            )
            holding_new = rebuild(new["code"], conn)
            holding_old = rebuild(row["code"], conn) if row["code"] != new["code"] else None
        except ValueError:
            raise
        except Exception as e:  # noqa: BLE001  UNIQUE 冲突等 → 转业务错误
            raise ValueError(f"修改失败: {e}")

    dates = {new["trade_time"][:10]}
    if row["trade_time"][:10] != new["trade_time"][:10]:
        dates.add(row["trade_time"][:10])
    # 涉及日 AI 打分失效并后台自动重打分
    for d in dates:
        _trigger_ai_daily(d)
    _rebuild_portfolio()
    return {"trade_id": trade_id, "holding": holding_new, "holding_old": holding_old}


def _trigger_ai_daily(score_date: str):
    """该日 AI 打分失效并后台自动重打分（尽力而为，失败不影响交易操作）。

    必须在写事务外调用：AI 打分自身会写 ai_daily_reports，事务内再开写连接会锁库。
    """
    try:
        from app.services.ai_scoring import maybe_auto_score_daily

        maybe_auto_score_daily(score_date)
    except Exception:
        pass


def _rebuild_portfolio():
    """开仓/清仓/改仓后重算组合综合 PE/PB 序列缓存（权重已变，历史序列需重建）。

    失败不影响交易操作（返回 None），日志里记录原因。
    """
    import logging

    try:
        from app.analysis.portfolio import rebuild_portfolio_series

        return rebuild_portfolio_series()
    except Exception as e:  # noqa: BLE001 组合序列数据不全不中断交易
        logging.getLogger("holdings").warning("组合序列重算失败：%s", e)
        return None


def _is_new_stock(code: str) -> bool:
    """该股是否首笔交易（此前无任何交易记录 → 需同步数据）。

    用交易记录而非缓存判定：清仓会删除该股缓存，若按缓存判"新股"会把旧股误判并重复同步。
    record_trade 事务已提交，故 trades 中该股仅当前这一笔即视为首笔。
    """
    from app.models.db import get_conn

    try:
        with get_conn() as c:
            row = c.execute("SELECT COUNT(*) AS n FROM trades WHERE code=?", (code,)).fetchone()
        return row["n"] <= 1
    except Exception:  # noqa: BLE001
        return False


def _sync_stock_data(code: str):
    """开仓新股后同步其全部数据（日K/财务/百度序列/分位）。失败不阻断交易录入。"""
    import logging

    try:
        from app.services.refresh import sync_stock_full

        sync_stock_full(code)
    except Exception as e:  # noqa: BLE001 数据同步失败不阻断录入
        logging.getLogger("holdings").warning("新股数据同步失败 %s：%s", code, e)


def list_trades(code: str | None = None) -> list[dict]:
    """交易流水（含股票名称）。"""
    with get_conn() as c:
        if code:
            rows = c.execute(
                """SELECT t.*, s.name FROM trades t LEFT JOIN stocks s ON t.code=s.code
                   WHERE t.code=? ORDER BY t.trade_time, t.id""", (code,)
            ).fetchall()
        else:
            rows = c.execute(
                """SELECT t.*, s.name FROM trades t LEFT JOIN stocks s ON t.code=s.code
                   ORDER BY t.trade_time, t.id"""
            ).fetchall()
        return [dict(r) for r in rows]


def backfill_trade_cny() -> int:
    """回填缺失 amount_cny 的港股交易（用当前汇率近似），并重建持仓人民币成本。

    用于汇率源此前不可用、导致港股成本人民币缺失的情况。返回回填笔数。
    """
    from app.services.fx import get_fx_rate_cny

    with get_conn() as c:
        rows = c.execute(
            """SELECT t.*, s.currency FROM trades t
               LEFT JOIN stocks s ON t.code=s.code
               WHERE COALESCE(s.currency,'CNY')<>'CNY' AND t.amount_cny IS NULL""",
        ).fetchall()
        n = 0
        for t in rows:
            rate = get_fx_rate_cny(t["currency"], t["trade_time"][:10])
            if rate:
                c.execute(
                    "UPDATE trades SET fx_rate=?, amount_cny=? WHERE id=?",
                    (round(rate, 6), round((t["amount"] or 0) * rate, 2), t["id"]),
                )
                n += 1
        codes = {t["code"] for t in rows}
    for code in codes:
        with get_conn() as c:
            rebuild(code, c)
    return n


def get_holdings(active_only: bool = True) -> list[dict]:
    """持仓列表（含股票名称、币种、人民币成本、累计分红）。"""
    with get_conn() as c:
        sql = (
            "SELECT h.*, s.name, s.tag, s.currency FROM holdings h LEFT JOIN stocks s ON h.code=s.code"
            + (" WHERE h.status='active'" if active_only else "")
            + " ORDER BY h.quantity DESC"
        )
        rows = [dict(r) for r in c.execute(sql).fetchall()]
        # 累计分红：adjust 且 is_dividend=1 的成本摊薄（自动 + 手动除权）
        div_totals = dict(
            c.execute(
                "SELECT code, COALESCE(SUM(-amount),0) FROM trades WHERE side='adjust' AND is_dividend=1 GROUP BY code"
            ).fetchall()
        )
    for r in rows:
        r["currency"] = r.get("currency") or "CNY"
        r["tag"] = r.get("tag") or get_instrument(r["code"]).tag
        # 港股不再复用 ETF 分类：is_etf 仅按场内基金代码/ETF 标签判定
        r["is_etf"] = get_instrument(r["code"]).is_etf or r["tag"] == "ETF"
        # 人民币成本口径：港股缺汇率时置 None（missing_fx）
        if r["avg_cost_cny"] is None and r["currency"] == "CNY":
            r["avg_cost_cny"] = r["avg_cost"]
        r["missing_fx"] = r["currency"] != "CNY" and r["avg_cost_cny"] is None
        # 该股累计已收分红（除权摊薄金额的绝对值）
        r["total_dividend"] = round(div_totals.get(r["code"], 0.0), 2)
    return rows


def set_tag(code: str, tag: str, name: str | None = None) -> str:
    """设置/更新个股标签（自动建 stocks 行）。"""
    tag = (tag or "").strip()
    if not tag:
        raise ValueError("标签不能为空")
    mkt = _market_of(get_instrument(code))
    with get_conn() as c:
        c.execute(
            """INSERT INTO stocks(code, name, market, tag) VALUES(?,?,?,?)
               ON CONFLICT(code) DO UPDATE SET tag=excluded.tag""",
            (code, name or code, mkt, tag),
        )
    # 标签变化影响组合画像与当日交易评分：组合 AI 报告靠画像哈希变 stale（各组合独立保留），今日重触发
    _trigger_ai_daily(datetime.now().strftime("%Y-%m-%d"))
    return tag


def init_holdings(items: list[dict]) -> list[dict]:
    """批量初始化持仓：每项 {code, name, price, quantity, fee?}。"""
    results = []
    for it in items:
        r = record_trade(
            code=it["code"], side="buy", price=it["price"], quantity=it["quantity"],
            fee=it.get("fee", 0.0), name=it.get("name"),
        )
        results.append(r)
    return results
