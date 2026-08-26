#!/usr/bin/env python3
"""parse_excel.py — 归一化三表，输出 clean JSON，供后续维度使用。含分池标记（债当现金）。"""
import openpyxl, json, pathlib, re, yaml
from datetime import datetime, date
from collections import defaultdict

CONFIG_PATH = pathlib.Path(__file__).parent.parent / "config.yaml"

ALIASES = {
    "code": ["代码","证券代码","股票代码"],
    "name": ["名称","证券名称","股票名称"],
    "amount": ["发生金额","成交金额","金额","成交额","持有金额"],
    "cat": ["交易类别","业务类型","操作"],
    "price": ["成交价格","价格","成交价"],
    "qty": ["成交数量","数量","股数"],
}

def load_pool_config():
    if not CONFIG_PATH.exists(): return None
    cfg = yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8"))
    return cfg.get("pool")

def classify_pool(code, name, pool_cfg):
    if not code: return "risk"
    code=str(code).strip()
    if pool_cfg:
        if code in pool_cfg.get("cash_codes", []): return "cash"
        for pre in pool_cfg.get("cash_prefixes", []):
            if code.startswith(pre): return "cash"
        for kw in pool_cfg.get("cash_names_contain", []):
            if kw and name and kw in str(name): return "cash"
    if code.startswith("511") or code.startswith("159") or code=="511620": return "cash"
    return "risk"

def normalize_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, date): return v.strftime("%Y-%m-%d")
    s=str(v).strip()
    m=re.match(r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", s)
    if m: return f"{int(m[1]):04d}-{int(m[2]):02d}-{int(m[3]):02d}"
    return s[:10] if len(s)>=10 else s

def find_header(ws, aliases):
    header_row=1
    headers=[c.value for c in ws[1]]
    if all(v is None for v in headers):
        headers=[c.value for c in ws[2]]; header_row=2
    col_map={}
    for idx, h in enumerate(headers, start=1):
        if h is None: continue
        hs=str(h).strip()
        for key, names in aliases.items():
            if hs in names: col_map[key]=idx
    return col_map, header_row

def parse(path, out_dir="cache"):
    pool_cfg = load_pool_config()
    wb=openpyxl.load_workbook(path, data_only=True)
    out=pathlib.Path(out_dir); out.mkdir(parents=True, exist_ok=True)
    result={}
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        col_map, hr = find_header(ws, ALIASES)
        # 持仓表特殊：固定列位，避免别名误判
        is_hold = sheet in ("持仓数据","持仓","在持")
        rows=[]
        for r in range(hr+1, ws.max_row+1):
            if is_hold:
                code=ws.cell(r,1).value
                name=ws.cell(r,2).value
                amt=ws.cell(r,3).value  # 持有金额
                # 持仓表没有 cat/price/qty 的交易语义，单独存
                if code is None and name is None: continue
                scode=str(code).strip() if code not in (None,"") else None
                rows.append({
                    "ds": None,
                    "code": scode,
                    "name": name,
                    "amount": amt or 0,
                    "holding_pnl": ws.cell(r,10).value or 0,
                    "weight": ws.cell(r,17).value or 0,
                    "holding_days": ws.cell(r,19).value or 0,
                    "cat": None,
                    "qty": ws.cell(r,18).value or 0,
                    "price": ws.cell(r,21).value or 0,
                    "amt": amt or 0,
                    "fee": 0,
                    "pool": classify_pool(scode, name, pool_cfg),
                    "sheet": sheet,
                })
            else:
                code=ws.cell(r, col_map.get("code",3)).value if "code" in col_map else ws.cell(r,3).value
                if code is None and all(ws.cell(r,c).value is None for c in range(1,5)): continue
                d=normalize_date(ws.cell(r,1).value)
                name = ws.cell(r, col_map.get("name",4)).value if "name" in col_map else ws.cell(r,4).value
                scode = str(code).strip() if code not in (None,"") else None
                rows.append({
                    "ds": d,
                    "code": scode,
                    "name": name,
                    "cat": ws.cell(r, col_map.get("cat",5)).value if "cat" in col_map else ws.cell(r,5).value,
                    "qty": ws.cell(r, col_map.get("qty",6)).value or 0,
                    "price": ws.cell(r, col_map.get("price",7)).value or 0,
                    "amt": ws.cell(r, col_map.get("amount",8)).value or 0,
                    "fee": ws.cell(r,10).value or 0,
                    "pool": classify_pool(scode, name, pool_cfg),
                    "sheet": sheet,
                })
        result[sheet]=rows
        pathlib.Path(out/f"{sheet}.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    all_trades=sum(result.values(), [])
    sec=[t for t in all_trades if t["cat"] in ("买入","卖出")]
    for t in sec: t["abs_amt"]=abs(t["amt"] or 0)
    pathlib.Path(out/"trades_clean.json").write_text(json.dumps(sec, ensure_ascii=False, indent=2), encoding="utf-8")
    risk=len([t for t in sec if t["pool"]=="risk"]); cash=len([t for t in sec if t["pool"]=="cash"])
    print(f"parsed {len(all_trades)} rows, sec {len(sec)} (risk {risk} / cash {cash}) -> {out}")
    return result

if __name__=="__main__":
    import sys
    parse(sys.argv[1], sys.argv[2] if len(sys.argv)>2 else "cache")
